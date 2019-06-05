from openpyxl import Workbook
from openpyxl import load_workbook

print('Opening workbook...')
wb = load_workbook('d:/python-execl/censuspopdata.xlsx')
ws = wb.active
print(wb.get_sheet_names())
print(wb.worksheets[0])
print(ws.max_row)
countyData = {}
print('Reading rows...')

for row in range(2,(ws.max_row + 1)):
    state= ws['B' + str(row)].value
    county = ws['C' + str(row)].value
    pop = ws['D' + str(row)].value

    countyData.setdefault(state,{})
    countyData[state].setdefault(county,{'tracts':0,'pop':0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)


