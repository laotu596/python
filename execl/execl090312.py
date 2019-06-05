from openpyxl import Workbook
from openpyxl import load_workbook

import datetime


wb = load_workbook('d:/python-execl/example.xlsx')
#print(type(wb))
ws = wb.get_sheet_names()
ws = wb.get_sheet_by_name('Sheet1')
c = ws['B1']
print(ws['A1'].value)
print(ws['B1'].value)
print('Row' + str(c.row) +', Column' + c.column + ' is ' + c.value )
print('Cell ' + c.coordinate + ' is ' + c.value)
print(ws['C1'].value)
print(ws.cell(row=1,column=2))
print(ws.cell(row=1,column=2).value)

for i in range(1,8,2):
    print(i,ws.cell(row=i,column=2).value)

print(tuple(ws['A1':'C3']))

for rowi in ws['A1':'C3']:
    for cellj in rowi:
        print(cellj.coordinate,cellj.value)
    print('---END OF ROW---')

#ws = wb.active
#wb.read_only
#wb.worksheets
#print(wb.properties)
#wb.encoding
#print(wb.get_sheet_names())
