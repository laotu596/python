from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

#openpyxl 中有三个不同层次的类，Workbook是对工作薄的抽象，Worksheet是对表格的抽象，Cell是对单元格的抽象
#操作Excel 的一般场景
#1、打开或者创建一个Excel需要创建一个Workbook对象
#2、获取一个表则需要先创建一个Workbook对象，然后使用该对象的方法来得到一个Worksheet对象
#3、如果要获取表中的数据，那么得到Worksheet对象以后再从中获取代表单元格的Cell对象
#Workbook对象
#一个Workbook对象代表一个Excel文档，因此再操作Excel之前，都应该先创建一个Workbook对象，对于创建一个新的Excel文档，直接进行Workbook类的调用即可
#对于一个已经存在的Excel文档，可以使用openpyxl模块的load_workbook函数进行去读
#一个Excel电子表格文档称为一个工作簿，一个工作簿保存在扩展名为.xlsx 的文件中；
# 每个工作簿可以包含多个表（也称为工作表）；
# 用户当前查看的表（或关闭Excel 前最后查看的表或Excel 在打开时出现的表），称为活动表。
#openpyxl 中主要得三个概念:Workbook(工作表)、Sheet(表页) 和Cell（挌）
#openpyxl 中主要得操作：打开Workbook,定位Sheet,操作Cell

#创建工作薄
wb = Workbook()                                                     #创建一个工作薄
ws = wb.active                                                      #获取工作得激活工作表

wb.save('文件名称.xlsx')                                          #保存创建好得工作薄

#打开已有得execl 文件
wb = load_workbook('文件名称.xlsx')
ws = wb.active
print(ws.max_row)                                                       #获取文件做大行数
print(ws.max_column)                                                    #获取文件最大列数

#workbook提供的方法如下
#get_sheet_names 获取所有表格的名称（新版已经不建议使用，通过Workbook的sheetnames属性即可获取）
#get_sheet_by_name 通过表格名称获取Worksheet对象（新版也不建议使用，通过Worksheet['表名']获取）
#get_active_sheet  获取活跃的表格（新版建议通过active属性获取）
#remove_sheet 删除一个表格
#create_sheet 创建一个表格
#copy_worksheet 在Workbook内拷贝表格

#sheet 操作
ws = wb.get_sheet_names                                              #得到sheet名字
ws = wb.create_sheet("Mysheet",0)                                 #创建表名称 sheet 插在了开始得位置，从0开始
ws = wb.create_sheet()                                              #创建表名称 sheet 插在了工作表末尾
ws.title = "hello"                                                 #更改sheet得名字
ws = wb.remove(my_sheet)                                            #删除某个工
#cell 操作
ws['A1'] = 42                                                      #在A1中插入内容
ws.append([1,2,3])                                                  #在第二行插入三个数字，占用三个单元格，即'A2=1','B2'=2 'C2'=3
ws['A2'] = datetime.datetime.now().strftime("%Y-%m-%d")           #在A2中插入了当前时间，把上一句得那个A2=1 挤掉了
row = [1,2,3,4]
ws.append(row)                                                       #添加一行
rows = [
    ["ID","data1","data2"],
    [2,40,20],
    [3.40.25],
    [4.40.30],
]
ws.append(rows)                                                     #添加多行
wb.save('d:/2.xlsx')                                               #保存工作薄

打印表中数据的几种方式

#按行读取：按A1,B1,C1顺序返回
for row in my_sheet.rows:
    for cell in row:
        print(cell.value)

#按列读取：按A1,A2,A3顺序返回
for column in my_sheet.columns:
    for cell in column:
        print(cell.value)


#获取矩形区间数据
for i in range(1,4):
    for j in range(1,3):
        print(my_sheet,cell(row=i,column=i))

#Worksheet 对象
#title 表格的标题
#dimensions 表格的大小 这里的大小是指含有数据的表格的大小，即 左上角的坐标，右下角的坐标
#max_row|min_row   表格的最大行
#max_column|min_column 表格的最大列、最小列
#rows 按行获取单元格（Cell对象）-生成器
#columns 按列获取单元格（Cell对象）- 生成器
#freeze_pans  冻结窗格
#values  按行获取表格的内容（数据） - 生成器

#常用Worksheet 方法如下
#iter_wors 按行获取所有单元格，内置属性有（min_row、max_row、min_col、max_col）
#iter_columns 按列获取所有的单元格
#append  在表格末尾添加数据
#merged_cell 合并多个单元格
#unmerged_cells 移除合并的单元格

#cell 对象
#row 单元格所在的行
#column 单元格所在的列
#value 单元格的值
#coordinate 单元格的坐标


