from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)
import datetime


wb = load_workbook('d:/python-execl/log.xlsx')
ws = wb.active

#print(ws.max_row)
#print(ws.max_column)

#print(ws.cell(3,2).value)
#for row in ws.rows:
#    for cell in row:
#        print(cell.value)

for row in ws.rows:
    for cell in row:
        print(cell.value,end='\t')
    print()
    ws.append()
#    print(row)
#for i in range(1,3):
#    for j in range(1,2):
#        print(ws.cell(row=i,column=j).value)
chart = AreaChart()
chart.title = "Nginx log"
chart.style = 13
chart.x_axis.title = 'date'
chart.y_axis.title = 'vistor'

cats = Reference(ws, min_col=1, min_row=1, max_row=17)
data = Reference(ws, min_col=1, min_row=1, max_col=3, max_row=17)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "A20")

wb.save("d:/python-execl/log4.xlsx")