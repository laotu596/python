from openpyxl import Workbook
import datetime
import time

#创建一个excel 文件，并写入不同类的内容
# -*- coding: utf-8 -*-
wb = Workbook()

ws = wb.active
ws['A1'] = 42
ws['B1'] = "你好"+"automation test"

ws.append([1,2,3])

#ws['A2'] = datetime.datetime.now()
#ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒",time.localtime())
ws['A2'] = datetime.datetime.now()    #写入一个当前时间
#写入一个自定义的时间格式
ws['A3'] =time.strftime('%Y{y}%m{m}%d{d} %H{h}%M{f}%S{s}').format(y='年',m='月',d='日',h='时',f='分',s='秒')
wb.save("d:\\demo.xlsx")
