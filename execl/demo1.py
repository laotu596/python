from openpyxl import Workbook


# -*- coding: utf-8 -*-
#创建sheet

wb = Workbook()

ws1 = wb.create_sheet('Mysheet')
ws1.title = "New Title"
ws2 = wb.create_sheet('Mysheet',0)
ws2.title = u"你好"

ws1.sheet_properties.tabColor= "1072BA"

print(wb.get_sheet_by_name(u"你好" ))
print(wb["New Title"])

print(wb.sheetnames)
for sheet_name in wb.sheetnames:
    print(sheet_name)

print("*"*50)

for sheet in wb:
    print(sheet.title)

wb["New Title"]["A1"]="zeke"
source = wb["New Title"]
target = wb.copy_worksheet(source)

wb.save("D:\\demo1.xlsx ")