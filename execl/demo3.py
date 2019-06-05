# -*- coding:utf-8 -*-

#批量操作单元格

from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("Mysheet")

ws["A1"]=1
ws["A2"]=2
ws["A3"]=3

ws["B1"]=4
ws["B2"]=5
ws["B3"]=6

ws["C1"]=7
ws["C2"]=8
ws["C3"]=9

print(ws["A"])
for cell in ws["A"]:
    print(cell.value)

print(ws["A:C"])
for column in ws["A:C"]:
    for cell in column:
        print(cell.value)

row_range = ws[1:3]
print(row_range)
for row in row_range:
    for cell in row:
        print(cell.value)

print("*"*50)
for row in ws.rows:
    print(row)

print("*"*50)
for col in ws.columns:
    print(col)

wb.save("d:\\python\demo3.xlsx")