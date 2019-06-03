# -*- coding:utf-8 -*-

from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet("Mysheet")

ws["A1"]=123.11
ws["B2"]="你好"

d = ws.cell(row=4,column=2,value=10)
print(ws['A1'].value)
print(ws['B2'].value)
print(d.value)

wb.save("d:\\python\\demo2.xlsx")