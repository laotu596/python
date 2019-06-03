import  pymysql
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter



conn = pymysql.connect(host='localhost',
                              user='shurli',
                              passwd='shurli@2018',
                              port=3306,
                              db="test"
                       )

cursor = conn.cursor()
sql = cursor.execute('select * from products')

datas = cursor.fetchall()
conn.commit()
#fields_complex = cursor.description
#print(fields_complex)
cursor.close()
conn.close()
#print(datas)


def write_excel_openpyxl(datas,filename):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = 'products'
    row0 = ['prod_id','vend_id','prod_name','prod_price','prod_desc']
#    print(len(row0))
    for k in range(len(row0)):
#        print(k)
        ws.cell(row=1,column=k+1).value=row0[k]
    for i in range(1,len(datas)+1):
        for j in range(1,len(row0)+1):
            ws.cell(row=i+1,column=j).value=datas[i-1][j-1]
    wb.save(filename = filename)
write_excel_openpyxl(datas,'d:\\demo3.xlsx')