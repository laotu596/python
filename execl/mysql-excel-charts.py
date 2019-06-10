import pymysql
from openpyxl.workbook import  Workbook
from openpyxl import load_workbook
from openpyxl.chart import Series,LineChart,Reference,BarChart

conn = pymysql.connect(host='localhost',
                              user='shurli',
                              passwd='shurli@2018',
                              port=3306,
                              db="test"
                       )

cursor = conn.cursor()
sql = cursor.execute('select * from orders')

datas = cursor.fetchall()
conn.commit()
cursor.close()
conn.close()

def write_excel_openpyxl(datas,filename):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = 'orders'
    row0 = ['order_num','order_date','cust_id']
#    print(len(row0))
    for k in range(len(row0)):
#        print(k)
        ws.cell(row=1,column=k+1).value=row0[k]
    for i in range(1,len(datas)+1):
        for j in range(1,len(row0)+1):
            ws.cell(row=i+1,column=j).value=datas[i-1][j-1]
    wb.save(filename = filename)
write_excel_openpyxl(datas,'d:\\python\\orders.xlsx')
#wb = Workbook()
#ws = wb.active

wb = load_workbook('D://python/orders.xlsx')
ws = wb.active
print(ws.max_row)
print(ws.max_column)

#从将表格中涉及的要画图的数据使用：Reference 创建一个对象
#比如：我选取 data = Reference(ws, min_col=5, min_row=4, max_col=10, max_row=4)
#参数含义：ws 一个活跃的sheet，数据来源。可以使用ws = wb.active 获取
#其他的就是指定这个表中的行列数据了：起始行、起始列、终止行、终止列

#通过传入Reference对象，创建一个Series对象

#创建一个Chart对象

#可选择的设置Chart对象的长（drawing.height）、宽（drawing.width）、坐标位置（drawing.top、drawing.left）。

#将Chart对象添加到Worksheet对象。
#定义类型图
chart = BarChart()
#定义Y轴
chart.y_axis.title = 'order_num'
#定义X 皱
chart.x_axis.title = 'cust_id'

refObj = Reference(ws,min_col=1,min_row=1,max_col=1,max_row=5)
seriesObj = Series(refObj,title='orders')


#seriesObj = Series(data,title='压力')
chart.append(seriesObj)
ws.add_chart(chart,"C8")
wb.save("D://python/mysql-execl-linechart2.xlsx")



