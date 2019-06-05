from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook('d:/python-execl/movie.xlsx')
ws = wb.active

print("数据表最大行数:", ws.max_row)
print("数据表最大列数:", ws.max_column)
print("Work Sheet Title:",ws.title)
print(ws.cell(row=1,column=1).value)
DATA = []

for row in ws.iter_rows(min_col=1,min_row=1,max_row=ws.max_row,max_col=ws.max_column):
    ROW = []
    for cell in row:
        ROW.append(cell.value)
        DATA.append(ROW)
print("表中读取到得数据:",end="")
print(DATA)